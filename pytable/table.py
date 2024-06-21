from typing import List, Literal
from pydantic import BaseModel
from io import StringIO, IOBase
from csv import DictWriter

try:
    # Optional
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    pass


class Table():
    rows: List[BaseModel] = []

    def __init__(self, model: BaseModel, initial: List[BaseModel] = None):
        self.model = model

        if initial:
            self.rows = initial

    def add(self, row: BaseModel) -> None:
        self.rows.append(row)

    def to_csv(self, delimiter: str = ',') -> StringIO:
        buffer = StringIO()

        csv = DictWriter(buffer, fieldnames=self._get_headers(), delimiter=delimiter)
        csv.writeheader()

        csv.fieldnames = list(self.model.model_fields.keys())

        for row in self.rows:
            csv.writerow(row.model_dump())

        return buffer

    def to_excel(self, sheet_name: str = 'Sheet1') -> Workbook:
        workbook = Workbook()
        sheet: Worksheet = workbook.create_sheet(sheet_name)

        sheet.append(self._get_headers())

        for row in self.rows:
            as_dict = row.model_dump(by_alias=True)
            sheet.append(list(as_dict.values()))

        if 'Sheet' in workbook.sheetnames and sheet_name != 'Sheet':
            del workbook['Sheet']

        return workbook
    
    def save(self, file: IOBase, format: Literal['csv', 'xlsx']) -> None:
        if format == 'csv':
            csv = self.to_csv(delimiter=',')
            file.write(csv.buffer.read())
        elif format == 'xlsx':
            excel = self.to_excel()
            excel.save(file)
        else:
            raise Exception(f'Unhandled format `{format}`')
    
    def _get_headers(self):
        """Get title for tables name from field's Title. if not set, use field's name on model instead"""
        return [(field.title if field.title else name) for name, field in self.model.model_fields.items()]
